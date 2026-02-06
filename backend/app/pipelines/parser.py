"""
RAIS Backend - Excel Parser Pipeline
Handles parsing of all 6 Excel file types with proper error handling
"""
from pathlib import Path
from typing import Optional
from datetime import datetime, date
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.models import FileType, DataSource


# ============================================================================
# FILE TYPE DETECTION
# ============================================================================

FILE_TYPE_PATTERNS = {
    FileType.PRODUCTION_CUMULATIVE: [
        r"yearly.*production.*commulative",
        r"production.*commulative.*\d{4}",
    ],
    FileType.CUMULATIVE: [
        r"^commulative.*\d{4}",
    ],
    FileType.ASSEMBLY: [
        r"assembly.*rejection.*report",
    ],
    FileType.VISUAL: [
        r"visual.*inspection.*report",
    ],
    FileType.INTEGRITY: [
        r"balloon.*valve.*integrity",
        r"integrity.*inspection",
    ],
    FileType.SHOPFLOOR: [
        r"shopfloor.*rejection.*report",
    ],
}


def detect_file_type(filename: str) -> FileType:
    """Detect file type from filename"""
    name_lower = filename.lower()
    
    for file_type, patterns in FILE_TYPE_PATTERNS.items():
        for pattern in patterns:
            if re.search(pattern, name_lower):
                return file_type
    
    return FileType.UNKNOWN


# ============================================================================
# EXCEL DATE CONVERSION
# ============================================================================

EXCEL_EPOCH = datetime(1899, 12, 30)


def excel_serial_to_date(serial: int | float) -> date | None:
    """Convert Excel serial date to Python date"""
    try:
        if pd.isna(serial):
            return None
        if isinstance(serial, (datetime, date)):
            return serial if isinstance(serial, date) else serial.date()
        if isinstance(serial, str):
            # Try parsing common date formats
            for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%B %Y", "%b %Y"]:
                try:
                    return datetime.strptime(serial.strip(), fmt).date()
                except ValueError:
                    continue
            return None
        
        serial = float(serial)
        if serial < 1:
            return None
        
        return (EXCEL_EPOCH + pd.Timedelta(days=serial)).date()
    except Exception:
        return None


# ============================================================================
# MERGED CELL HANDLING
# ============================================================================

def unmerge_and_fill(workbook) -> None:
    """Unmerge all cells and fill with top-left value"""
    for sheet in workbook.worksheets:
        # Get list of merged ranges (copy to avoid modification during iteration)
        merged_ranges = list(sheet.merged_cells.ranges)
        
        for merged_range in merged_ranges:
            # Get top-left cell value
            top_left = sheet.cell(merged_range.min_row, merged_range.min_col)
            value = top_left.value
            
            # Unmerge
            sheet.unmerge_cells(str(merged_range))
            
            # Fill all cells with the value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    sheet.cell(row, col, value)


# ============================================================================
# HEADER DETECTION
# ============================================================================

HEADER_KEYWORDS = {
    "s.no", "s.no.", "sl.no", "date", "month", "year",
    "production", "dispatch", "received", "inspected", "accepted", "rejected",
    "qty", "quantity", "total", "percentage", "%", "defect",
    "batch", "lot", "item", "product", "code", "remarks", "result"
}


def score_row_as_header(row: list) -> float:
    """Score a row to determine if it's a header row"""
    if not row:
        return 0
    
    score = 0
    non_empty = 0
    
    for cell in row:
        if cell is None or (isinstance(cell, str) and not cell.strip()):
            continue
        
        non_empty += 1
        cell_str = str(cell).lower().strip()
        
        # Check for header keywords
        for keyword in HEADER_KEYWORDS:
            if keyword in cell_str:
                score += 10
                break
        
        # Prefer strings over numbers for headers
        if isinstance(cell, str):
            score += 2
        elif isinstance(cell, (int, float)):
            score -= 1  # Penalize numeric values in headers
    
    # Normalize by non-empty cells
    if non_empty > 0:
        score = score / non_empty * min(non_empty, 10)
    
    return score


def find_header_row(sheet, max_rows: int = 20) -> int:
    """Find the most likely header row in a sheet"""
    best_score = 0
    best_row = 0
    
    for row_idx in range(1, min(max_rows + 1, sheet.max_row + 1)):
        row_values = [sheet.cell(row_idx, col).value for col in range(1, sheet.max_column + 1)]
        score = score_row_as_header(row_values)
        
        if score > best_score:
            best_score = score
            best_row = row_idx
    
    return best_row


# ============================================================================
# MAIN PARSER
# ============================================================================

class ParsedSheet:
    """Represents a parsed sheet"""
    def __init__(
        self,
        name: str,
        headers: list[str],
        data: list[dict],
        header_row: int,
        file_name: str
    ):
        self.name = name
        self.headers = headers
        self.data = data
        self.header_row = header_row
        self.file_name = file_name
        self.row_count = len(data)


class ParseResult:
    """Result of parsing an Excel file"""
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.file_name = Path(file_path).name
        self.file_type = detect_file_type(self.file_name)
        self.sheets: list[ParsedSheet] = []
        self.errors: list[str] = []
        self.success = False


def clean_value(value):
    """Clean a cell value, handling special cases"""
    if value is None:
        return None
    
    # Handle error values
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in ["#DIV/0!", "#N/A", "#VALUE!", "#REF!", "#NAME?"]:
            return None
        if stripped == "":
            return None
        return stripped
    
    # Handle negative numbers (take absolute value with flag)
    if isinstance(value, (int, float)):
        if pd.isna(value):
            return None
        # Keep negative for now, validator will handle
        return value
    
    return value


def parse_sheet(sheet, file_name: str) -> ParsedSheet | None:
    """Parse a single sheet into structured data"""
    if sheet.max_row < 2:
        return None
    
    # Find header row
    header_row_idx = find_header_row(sheet)
    if header_row_idx == 0:
        header_row_idx = 1
    
    # Extract headers
    headers = []
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(header_row_idx, col).value
        if value:
            headers.append(str(value).strip().upper().replace("\n", " "))
        else:
            headers.append(f"COL_{col}")
    
    # Remove duplicate empty columns at the end
    while headers and headers[-1].startswith("COL_"):
        # Check if the column has any data
        has_data = False
        col_idx = len(headers)
        for row_idx in range(header_row_idx + 1, min(header_row_idx + 10, sheet.max_row + 1)):
            if sheet.cell(row_idx, col_idx).value is not None:
                has_data = True
                break
        if not has_data:
            headers.pop()
        else:
            break
    
    # Extract data rows
    data = []
    empty_row_count = 0
    
    for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
        row_data = {}
        has_value = False
        
        for col_idx, header in enumerate(headers, 1):
            value = clean_value(sheet.cell(row_idx, col_idx).value)
            row_data[header] = value
            if value is not None:
                has_value = True
        
        if has_value:
            row_data["_source_row"] = row_idx
            data.append(row_data)
            empty_row_count = 0
        else:
            empty_row_count += 1
            # Stop if too many empty rows
            if empty_row_count > 10:
                break
    
    if not data:
        return None
    
    return ParsedSheet(
        name=sheet.title,
        headers=headers,
        data=data,
        header_row=header_row_idx,
        file_name=file_name
    )


def parse_excel_file(file_path: str) -> ParseResult:
    """
    Parse an Excel file with full error handling.
    Handles merged cells, formulas, date serials, and special values.
    """
    result = ParseResult(file_path)
    
    try:
        # Load workbook with data_only to evaluate formulas
        wb = load_workbook(file_path, data_only=True)
        
        # Unmerge and fill cells
        unmerge_and_fill(wb)
        
        # Parse each sheet
        for sheet in wb.worksheets:
            # Skip sheets with specific patterns (often summary/chart sheets)
            sheet_name_lower = sheet.title.lower()
            if any(skip in sheet_name_lower for skip in ["chart", "graph", "summary"]):
                continue
            
            parsed = parse_sheet(sheet, result.file_name)
            if parsed:
                result.sheets.append(parsed)
        
        if result.sheets:
            result.success = True
        else:
            result.errors.append("No valid data sheets found")
        
        wb.close()
        
    except Exception as e:
        result.errors.append(f"Failed to parse file: {str(e)}")
    
    return result


def parse_multiple_files(file_paths: list[str]) -> dict[FileType, list[ParseResult]]:
    """Parse multiple Excel files and group by type"""
    results: dict[FileType, list[ParseResult]] = {ft: [] for ft in FileType}
    
    for file_path in file_paths:
        result = parse_excel_file(file_path)
        results[result.file_type].append(result)
    
    return results
